"""Dishonest Query Task - End-to-End Verifiable (strict & concise).

Category: law
Task tag: dishonest_query

Scoring (as requested):
- Treat EACH ROW in ground_truth.xlsx as ONE scoring item ("test point").
- A ground-truth row is correct if the model output contains a matching row,
  while IGNORING these three columns: 出生年份, 身份证归属地, 居住地.
- Full score: 1.0
- Each missing ground-truth row deducts 0.1
- Minimum: 0.0

Strictness:
- If the output file is missing, or output header/sheet format is not EXACT -> score = 0.0.
- Only .xlsx is accepted.

Normalization in matching:
- Ignore all whitespace (spaces, newlines, tabs, etc.).
- Treat Chinese/English parentheses as equivalent: （） == ().
- Support duplicates correctly: each ground-truth row needs its own match (multiset matching).
"""

import logging
import math
import re
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Dict, List, Optional, Sequence, Tuple

import cua_bench as cb
from openpyxl import load_workbook
from tasks.common_config import GeneralTaskConfig

logger = logging.getLogger(__name__)


# -----------------------------
# Task Config
# -----------------------------
@dataclass
class TaskConfig(GeneralTaskConfig):
    TASK_TAG: str = "dishonest_query"
    TASK_CATEGORY: str = "law"
    OUTPUT_FILENAME: str = "dishonest_query_output.xlsx"

    @property
    def problem_url(self) -> str:
        return rf"{self.task_dir}\input\problem.xlsx"

    @property
    def ground_truth_url(self) -> str:
        return rf"{self.reference_dir}\ground_truth.xlsx"

    @property
    def output_url(self) -> str:
        return rf"{self.remote_output_dir}\{self.OUTPUT_FILENAME}"

    @property
    def task_description(self) -> str:
        # The tested model cannot access ground_truth.xlsx, so the output format is specified explicitly.
        return f"""
Goal
You are given an Excel file of queries. For each query, find ALL matching "被执行人 / 失信被执行人" records from the Internet
and write them to a single Excel output file.

Input
- File: {self.problem_url}
- Each row is ONE test case (query).
- Columns (Chinese headers in the file):
  - 姓名 (required)
  - 出生年份 (optional; NaN/empty means no restriction)
  - 身份证归属地 (optional; NaN/empty means no restriction)
  - 居住地 (optional; NaN/empty means no restriction)

Output (STRICT; otherwise score = 0.0)
- Save EXACTLY one Excel file to: {self.output_url}
- Only one sheet (the first/active sheet).
- Row 1 must be the header with EXACTLY 11 columns in EXACT order:

  1) 姓名
  2) 出生年份
  3) 身份证归属地
  4) 居住地
  5) 身份证号
  6) 执行依据文号
  7) 案号
  8) 做出执行依据单位
  9) 生效法律文书确定的义务
  10) 被执行人的履行情况
  11) 失信被执行人行为具体情形

- Each record must take ONE row (starting from Row 2).
- If you cannot obtain 出生年份 / 身份证归属地 / 居住地 from sources, you may leave those cells empty.
- Do NOT merge cells.

Notes
- If a query matches multiple records (even hundreds or thousands), you MUST output all of them.
- You do NOT need to group by query; just output all records row-by-row.
- Evaluation is done by matching 身份证号. So, if you really cannot find 身份证号 for a specific record, you may leave it empty.

The task is considered successful if:
- The output file is well-formed and follows the strict format above.
- The output file contains specific records (i.e., test points) for each query. All the columns of the test points are guaranteed to be found online.
"""

    def to_metadata(self) -> dict:
        md = super().to_metadata()
        md.update(
            {
                "problem_url": self.problem_url,
                "ground_truth_url": self.ground_truth_url,
                "output_url": self.output_url,
                "output_filename": self.OUTPUT_FILENAME,
            }
        )
        return md


config = TaskConfig()


@cb.tasks_config(split="train")
def load():
    return [
        cb.Task(
            description=config.task_description,
            metadata=config.to_metadata(),
            computer={"provider": "computer", "setup_config": {"os_type": config.OS_TYPE}},
        )
    ]


@cb.setup_task(split="train")
async def start(task_cfg, session: cb.DesktopSession):
    """Initialize: clean output dir; open input file (optional)."""
    logger.info(f"Setting up task: {config.TASK_TAG}")
    try:
        await session.remove_file(task_cfg.metadata["remote_output_dir"])
        await session.makedirs(task_cfg.metadata["remote_output_dir"])
        await session.run_file(task_cfg.metadata["problem_url"])
    except Exception as e:
        logger.warning(f"Failed to setup tasks {config.TASK_TAG} via session: {e}")


# -----------------------------
# Strict Parsing + Matching
# -----------------------------
OUTPUT_HEADER: Sequence[str] = (
    "姓名",
    "出生年份",
    "身份证归属地",
    "居住地",
    "身份证号",
    "执行依据文号",
    "案号",
    "做出执行依据单位",
    "生效法律文书确定的义务",
    "被执行人的履行情况",
    "失信被执行人行为具体情形",
)

# Matching ignores these 3 columns.
KEY_COLS: Sequence[str] = (
    "姓名",
    "身份证号",
    "执行依据文号",
    "案号",
    "做出执行依据单位",
    "生效法律文书确定的义务",
    "被执行人的履行情况",
    "失信被执行人行为具体情形",
)


def _is_blank(x: Any) -> bool:
    if x is None:
        return True
    if isinstance(x, float) and math.isnan(x):
        return True
    if isinstance(x, str) and x.strip() == "":
        return True
    if isinstance(x, str) and x.strip().lower() in {"nan", "none", "null"}:
        return True
    return False


def _strip_trailing_blanks(row: Sequence[Any]) -> List[Any]:
    r = list(row)
    while r and _is_blank(r[-1]):
        r.pop()
    return r


def _read_sheet_rows(xlsx_bytes: bytes) -> List[List[Any]]:
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _first_nonempty_idx(rows: List[List[Any]]) -> Optional[int]:
    for i, r in enumerate(rows):
        if r and any(not _is_blank(c) for c in r):
            return i
    return None


def _canon_text(x: Any) -> str:
    """Ignore all whitespace; Chinese/English parentheses equivalent."""
    if _is_blank(x):
        return ""
    s = str(x)
    s = s.replace("（", "(").replace("）", ")")
    s = re.sub(r"\s+", "", s)
    return s


def _strict_header_ok(row: Sequence[Any], expected: Sequence[str]) -> bool:
    r = _strip_trailing_blanks(row)
    if len(r) != len(expected):
        return False
    return tuple(str(c).strip() if not _is_blank(c) else "" for c in r) == tuple(expected)


def _strict_parse_table(xlsx_bytes: bytes) -> Optional[List[Dict[str, Any]]]:
    """Parse an xlsx table with the exact OUTPUT_HEADER. Return None if header/shape is invalid."""
    rows = _read_sheet_rows(xlsx_bytes)
    i0 = _first_nonempty_idx(rows)
    if i0 is None:
        return None
    if not _strict_header_ok(rows[i0], OUTPUT_HEADER):
        return None

    out: List[Dict[str, Any]] = []
    for r in rows[i0 + 1 :]:
        rr = _strip_trailing_blanks(r)
        if not rr or all(_is_blank(c) for c in rr):
            continue
        # Strict: no extra non-empty columns beyond OUTPUT_HEADER
        if len(rr) > len(OUTPUT_HEADER) and any(not _is_blank(c) for c in rr[len(OUTPUT_HEADER) :]):
            return None

        d: Dict[str, Any] = {}
        for j, h in enumerate(OUTPUT_HEADER):
            d[h] = rr[j] if j < len(rr) else None
        out.append(d)
    return out


def _key(row: Dict[str, Any]) -> Tuple[str, ...]:
    return tuple(_canon_text(row.get(c)) for c in KEY_COLS)


# -----------------------------
# Evaluation (no VLM; no EvaluationContext)
# -----------------------------
@cb.evaluate_task(split="train")
async def evaluate(task_cfg, session: cb.DesktopSession) -> list[float]:
    try:
        gt_bytes = await session.read_bytes(task_cfg.metadata["ground_truth_url"])

        try:
            out_bytes = await session.read_bytes(task_cfg.metadata["output_url"])
        except Exception:
            return [0.0]

        gt_rows = _strict_parse_table(gt_bytes)
        out_rows = _strict_parse_table(out_bytes)
        if gt_rows is None or out_rows is None:
            return [0.0]

        # Multiset matching: handle duplicates correctly.
        out_counter = Counter(_key(r) for r in out_rows)

        fail = 0
        for gr in gt_rows:
            k = _key(gr)
            if out_counter.get(k, 0) > 0:
                out_counter[k] -= 1
            else:
                logger.info(f"Missing ground-truth row: {gr}")
                fail += 1

        score = max(1.0 - 0.1 * fail, 0.0)
        return [score]

    except Exception as e:
        logger.error(f"Evaluation error: {e}")
        return [0.0]
